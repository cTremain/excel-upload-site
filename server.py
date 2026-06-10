import json
import mimetypes
import re
import uuid
from datetime import datetime
from email.parser import BytesParser
from email.policy import default
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


ROOT = Path(__file__).resolve().parent
PUBLIC_DIR = ROOT / "public"
UPLOAD_DIR = ROOT / "uploads"
ORIGINALS_DIR = UPLOAD_DIR / "originals"
EDITED_DIR = UPLOAD_DIR / "edited"
MAX_UPLOAD_BYTES = 10 * 1024 * 1024
PORT = 3000


class ExcelWebsiteHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(PUBLIC_DIR), **kwargs)

    def do_POST(self):
        if self.path == "/api/upload":
            self.handle_upload()
            return

        if self.path == "/api/download":
            self.handle_download()
            return

        self.send_json({"error": "Not found."}, HTTPStatus.NOT_FOUND)

    def handle_upload(self):
        content_length = int(self.headers.get("Content-Length", "0"))
        if content_length > MAX_UPLOAD_BYTES + 1024 * 1024:
            self.send_json({"error": "The file is too large. Maximum size is 10 MB."}, HTTPStatus.BAD_REQUEST)
            return

        uploaded = parse_multipart_upload(self.headers, self.rfile.read(content_length), "workbook")
        if uploaded is None:
            self.send_json({"error": "No Excel file was uploaded."}, HTTPStatus.BAD_REQUEST)
            return

        original_name = safe_filename(uploaded["filename"])
        if Path(original_name).suffix.lower() != ".xlsx":
            self.send_json({"error": "Only .xlsx files are supported."}, HTTPStatus.BAD_REQUEST)
            return

        ensure_storage()
        file_id = str(uuid.uuid4())
        stored_path = ORIGINALS_DIR / f"{file_id}-{original_name}"

        with stored_path.open("wb") as output:
            output.write(uploaded["content"])

        if stored_path.stat().st_size > MAX_UPLOAD_BYTES:
            stored_path.unlink(missing_ok=True)
            self.send_json({"error": "The file is too large. Maximum size is 10 MB."}, HTTPStatus.BAD_REQUEST)
            return

        try:
            workbook = load_workbook(stored_path, data_only=False)
            worksheet = workbook.worksheets[0]
            payload = {
                "fileId": file_id,
                "filename": original_name,
                "worksheet": serialize_worksheet(worksheet),
                "limits": {"maxUploadMb": 10},
            }
            self.send_json(payload)
        except Exception as error:
            stored_path.unlink(missing_ok=True)
            self.send_json({"error": f"Could not read the workbook: {error}"}, HTTPStatus.BAD_REQUEST)

    def handle_download(self):
        try:
            content_length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(content_length).decode("utf-8"))
        except Exception:
            self.send_json({"error": "Invalid download request."}, HTTPStatus.BAD_REQUEST)
            return

        file_id = payload.get("fileId")
        cells = payload.get("cells")
        if not file_id or not isinstance(cells, list):
            self.send_json({"error": "A fileId and cells array are required."}, HTTPStatus.BAD_REQUEST)
            return

        original_path = find_original_file(file_id)
        if original_path is None:
            self.send_json({"error": "Uploaded workbook was not found."}, HTTPStatus.NOT_FOUND)
            return

        workbook = load_workbook(original_path, data_only=False)
        worksheet = workbook.worksheets[0]

        for edit in cells:
            row = edit.get("row")
            col = edit.get("col")
            if not isinstance(row, int) or not isinstance(col, int):
                continue
            worksheet.cell(row=row, column=col).value = parse_cell_value(edit.get("value", ""))

        ensure_storage()
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        edited_path = EDITED_DIR / f"{file_id}-{timestamp}-edited.xlsx"
        workbook.save(edited_path)
        self.send_file(edited_path, f"edited-{original_path.name.replace(file_id + '-', '')}")

    def send_json(self, payload, status=HTTPStatus.OK):
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_file(self, path, download_name):
        content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        data = path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f'attachment; filename="{download_name}"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


def ensure_storage():
    ORIGINALS_DIR.mkdir(parents=True, exist_ok=True)
    EDITED_DIR.mkdir(parents=True, exist_ok=True)


def safe_filename(filename):
    name = Path(filename).name
    return re.sub(r"[^a-zA-Z0-9._-]", "_", name)


def parse_multipart_upload(headers, body, field_name):
    content_type = headers.get("Content-Type")
    if not content_type or "multipart/form-data" not in content_type:
        return None

    message_bytes = (
        f"Content-Type: {content_type}\r\n"
        "MIME-Version: 1.0\r\n\r\n"
    ).encode("utf-8") + body
    message = BytesParser(policy=default).parsebytes(message_bytes)

    for part in message.iter_parts():
        if part.get_param("name", header="content-disposition") != field_name:
            continue

        filename = part.get_filename()
        if not filename:
            return None

        return {
            "filename": filename,
            "content": part.get_payload(decode=True) or b"",
        }

    return None


def find_original_file(file_id):
    if not ORIGINALS_DIR.exists():
        return None
    matches = list(ORIGINALS_DIR.glob(f"{file_id}-*.xlsx"))
    return matches[0] if matches else None


def serialize_worksheet(worksheet):
    row_count = max(worksheet.max_row or 1, 20)
    col_count = max(worksheet.max_column or 1, 8)
    merges = [serialize_merge(str(item)) for item in worksheet.merged_cells.ranges]
    merge_children = set()

    for merge in merges:
        for row in range(merge["top"], merge["bottom"] + 1):
            for col in range(merge["left"], merge["right"] + 1):
                if row != merge["top"] or col != merge["left"]:
                    merge_children.add((row, col))

    columns = []
    for col in range(1, col_count + 1):
        letter = get_column_letter(col)
        width = worksheet.column_dimensions[letter].width or 12
        columns.append({"index": col, "letter": letter, "width": round(width * 8)})

    rows = []
    for row_index in range(1, row_count + 1):
        row_dimension = worksheet.row_dimensions[row_index]
        cells = []

        for col_index in range(1, col_count + 1):
            cell = worksheet.cell(row=row_index, column=col_index)
            merge = next((item for item in merges if item["top"] == row_index and item["left"] == col_index), None)
            cells.append(
                {
                    "row": row_index,
                    "col": col_index,
                    "address": cell.coordinate,
                    "value": cell_to_display_value(cell),
                    "style": style_to_css(cell),
                    "hiddenByMerge": (row_index, col_index) in merge_children,
                    "rowSpan": merge["bottom"] - merge["top"] + 1 if merge else 1,
                    "colSpan": merge["right"] - merge["left"] + 1 if merge else 1,
                }
            )

        rows.append({"index": row_index, "height": round(row_dimension.height or 24), "cells": cells})

    return {"name": worksheet.title, "rows": rows, "columns": columns, "merges": merges}


def serialize_merge(range_text):
    left, top, right, bottom = range_boundaries(range_text)
    return {"top": top, "left": left, "bottom": bottom, "right": right}


def cell_to_display_value(cell):
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return str(value)


def parse_cell_value(value):
    text = "" if value is None else str(value)
    if text.strip() == "":
        return None
    if text.startswith("=") and len(text) > 1:
        return text
    if re.fullmatch(r"-?\d+(\.\d+)?", text.strip()):
        return float(text) if "." in text else int(text)
    return text


def style_to_css(cell):
    style = {}

    if cell.font:
        if cell.font.bold:
            style["fontWeight"] = "700"
        if cell.font.italic:
            style["fontStyle"] = "italic"
        color = color_to_css(cell.font.color)
        if color:
            style["color"] = color

    if cell.fill and cell.fill.fill_type in {"solid", "darkGrid", "darkTrellis"}:
        color = color_to_css(cell.fill.fgColor)
        if color:
            style["backgroundColor"] = color

    if cell.alignment:
        if cell.alignment.horizontal:
            style["textAlign"] = cell.alignment.horizontal
        if cell.alignment.vertical:
            style["verticalAlign"] = cell.alignment.vertical

    return style


def color_to_css(color):
    if not color or color.type != "rgb" or not color.rgb:
        return None
    rgb = color.rgb
    if len(rgb) == 8:
        return f"#{rgb[2:]}"
    if len(rgb) == 6:
        return f"#{rgb}"
    return None


if __name__ == "__main__":
    ensure_storage()
    server = ThreadingHTTPServer(("", PORT), ExcelWebsiteHandler)
    print(f"Excel website running at http://localhost:{PORT}")
    server.serve_forever()
